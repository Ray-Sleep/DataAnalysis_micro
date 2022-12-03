from openpyxl import load_workbook

# 使用 openpyxl 模块读取 "出货统计表.xlsx" 中的数据
workbook = load_workbook('出货统计表.xlsx')
# 读取工作本中 名为 Sheet1 的工作表
worksheet = workbook['Sheet1']

# 使用for循环 遍历工作表中的数据，按出货日期顺序，分类存放到字典。
data = {}
for row in range(2, worksheet.max_row + 1):
    date = worksheet['B' + str(row)].value.date()
    customer = worksheet['C' + str(row)].value
    product = worksheet['D' + str(row)].value
    number = worksheet['E' + str(row)].value
    model = worksheet['G' + str(row)].value
    # 分别为 客户名 、 产品名称 、 出货数量 、 规格型号
    info_list = [customer, product, number, model]
    # 表示将出货日期作为键，不覆盖原键，而是将其值添加到后面的空列表，形成嵌套列表
    data.setdefault(date, [])
    data[date].append(info_list)

# for循环遍历字典 的 键、值，并打印
for key, value in data.items():
    print(key, value)

# 打开工作本"出货清单模板.xlsx"
workbook_day = load_workbook('出货清单模板.xlsx')
# 读取其中的工作表"出货清单模板"
worksheet_day = workbook_day['出货清单模板']


for date in data.keys():
    # 使用 copy_worksheet() 函数复制工作表后赋给 worksheet_new
    worksheet_new = workbook_day.copy_worksheet(worksheet_day)
    # 使用出货日期中的月和日重命名 工作表
    worksheet_new.title = str(date)[-5:]
    # 将出货日期写入 第2行 和 第5列 的单元格 （E2）
    worksheet_new.cell(row=2, column=5).value = date

    # 从第4行开始逐行填写出货记录
    i = 4
    for product in data[date]:
        worksheet_new.cell(row=i, column=2).value = product[0]
        worksheet_new.cell(row=i, column=3).value = product[1]
        worksheet_new.cell(row=i, column=4).value = product[2]
        worksheet_new.cell(row=i, column=5).value = product[3]
        i += 1

workbook_day.save('产品出货清单.xlsx')
